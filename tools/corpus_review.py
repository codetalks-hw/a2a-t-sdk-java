#!/usr/bin/env python3
"""Business review support for the A2A-T negotiation test corpus (design doc-local/fromtext-validate-test-design.md §8.7, Q19).

The corpus itself is the single-business-domain private-line complaint diagnosis corpus (Q20-Q23): every case,
scenario and golden fixture speaks the complaint-diagnosis five-step closed loop, and the dictionaries below
translate the corpus facts into that business wording (task API actions of A2ATClient/A2ATServer, task parameter
names, complaint-diagnosis business domains).

Review material is a *projection* of the corpus: generated from negotiation-cases/ JSON, never hand-maintained.

Two modes:

  export                          generate docs-local/review/corpus-review.html (self-contained dashboard +
                                  per-domain case cards + scenario ping-pong tables) and
                                  docs-local/review/corpus-review.xlsx (one row per case, last two columns are
                                  blank annotation columns for the business expert)
  collect --marked <file>         read the annotated xlsx (or a csv fallback) back, emit
                                  docs-local/review/review-findings.md (findings mapped to exact JSON file + line)
                                  and docs-local/review/review-status.json (per-case review status)

xlsx output requires `openpyxl` (pip install openpyxl). Without it, export degrades gracefully:
a warning is printed and only the HTML report is produced; collect then accepts a .csv marked file
(export also writes docs-local/review/corpus-review.csv alongside the xlsx so the closed loop works
in degraded environments).
"""

from __future__ import annotations

import argparse
import csv
import html
import io
import json
import sys
from datetime import datetime
from pathlib import Path

REPO_ROOT = Path(__file__).resolve().parent.parent
CORPUS_ROOT = REPO_ROOT / "a2a-t-corpus" / "src" / "test" / "resources" / "negotiation-cases"
REVIEW_DIR = REPO_ROOT / "docs-local" / "review"
HTML_PATH = REVIEW_DIR / "corpus-review.html"
XLSX_PATH = REVIEW_DIR / "corpus-review.xlsx"
CSV_PATH = REVIEW_DIR / "corpus-review.csv"
STATUS_PATH = REVIEW_DIR / "review-status.json"
FINDINGS_PATH = REVIEW_DIR / "review-findings.md"

CASE_FAMILIES = ("from-text", "from-data", "validate")
SCENARIO_FAMILY = "scenarios"

# --- business-language dictionaries (Q19: experts see pure business wording, never raw codes) -------------

API_ACTION_ZH = {
    "generateProposeFromText": "从自然语言生成提议消息",
    "generateAcceptFromText": "从自然语言生成接受消息",
    "generateRejectFromText": "从自然语言生成拒绝消息",
    "generateAbortFromText": "从自然语言生成终止消息",
    "generateProposeFromData": "从结构化数据生成提议消息",
    "generateAcceptFromData": "从结构化数据生成接受消息",
    "generateRejectFromData": "从结构化数据生成拒绝消息",
    "generateAbortFromData": "从结构化数据生成终止消息",
    "validateProposePromptAndDataFilling": "校验提议消息并提取参数",
    "validateAcceptPromptAndDataFilling": "校验接受消息并提取参数",
    "validateRejectPromptAndDataFilling": "校验拒绝消息并提取参数",
    "validateAbortPromptAndDataFilling": "校验终止消息并提取参数",
    # the three closed-loop task APIs (Q20-Q23): A2ATClient facade for generation, A2ATServer facade for validation
    "generateTaskPromptFromText": "工作台从投诉文本生成任务报文（A2ATClient）",
    "generateTaskPromptFromDataWithSchema": "工作台从结构化数据与参数模式生成任务报文（A2ATClient）",
    "validateTaskPromptAndDataFilling": "OMC 校验对端任务报文并提取任务参数（A2ATServer）",
}

# Task parameter names of the private-line complaint diagnosis (schema biz.complaint.params), for business wording.
TASK_PARAM_ZH = {
    "accessPort": "接入端口名称",
    "bizScenario": "投诉分类",
    "faultTime": "问题发生时间",
    "eventSerialNo": "OSS侧事件流水号",
}

# Shared schema references of the corpus, for business wording.
SCHEMA_ZH = {
    "biz.complaint.params": "专线投诉诊断业务参数模式（接入端口名称/投诉分类必填）",
}

ERROR_CODE_ZH = {
    "negotiation.invalid_input": "协商输入无效（如消息不是协商消息、缺少协商上下文）",
    "negotiation.field_missing": "协商消息缺少必填字段",
    "negotiation.content_extract_failed": "无法从文本提取协商内容",
    "negotiation.content_invalid": "from-data 协商内容字段无效（如空列表、空白描述）",
    "negotiation.conclusion_mismatch": "报文结论与该方法预期不符",
    "negotiation.semantic_rejected": "语义校验判定不通过",
    "negotiation.rule_violation": "协商报文违反校验规则（未知码兜底，含轮次超限、标识格式非法）",
    "negotiation.invalid_context_id": "协商上下文标识不是合法 UUID",
    "negotiation.round_exceeded": "协商轮次超过上限",
    "negotiation.type_mismatch": "报文内容与声明的模板类型不符",
    "template.not_found": "协商模板不存在或不可用",
    "llm.invocation_failed": "大模型调用失败（传输/网络层）",
    "llm.response_invalid": "大模型返回内容不符合要求（响应契约违反）",
}

EXCEPTION_ZH = {
    "NegotiationGenerationException": "协商生成失败",
    "NegotiationParamExtractionException": "协商参数提取失败",
    "NullPointerException": "空指针异常（编程错误）",
    "IllegalArgumentException": "非法参数异常（编程错误）",
    "RuntimeException": "运行时异常（编程错误）",
}

FAIL_MODE_ZH = {
    "non-json": "非法 JSON 数据",
    "null-response": "空响应",
    "blank-content": "空白内容",
    "llm-error": "基础设施错误",
    "assertion": "断言失败（不该被调用时触发）",
}

INJECT_ZH = {
    "failingTemplateLoader": "注入不可用的模板加载器（模拟模板缺失）",
    "failingSemanticValidator": "注入不可用的语义校验器（模拟校验器故障）",
}

TEMPLATE_URI_ZH = {
    # complaint-diagnosis business domains (Q20/Q23): every corpus record belongs to the private-line complaint
    # diagnosis; the domain labels carry its closed-loop semantics instead of bare negotiation type names
    "network-layer": "任务闭环（专线投诉诊断）",
    "information-negotiation": "信息协商（缺参要数）",
    "target-negotiation": "目标协商（修复目标确认）",
    "feasibility-negotiation": "可行性协商（修复方案确认）",
    "common": "通用（终止）",
}
PHASE_ZH = {
    "propose": "提议",
    "accept-reject": "接受/拒绝",
    "abort": "终止",
    "private-line-complaint": "任务报文",
}
TERMINAL_ZH = {"accept": "接受（协商达成）", "reject": "拒绝（协商终止）", "abort": "终止", "exhausted": "轮次预算耗尽"}

DOMAIN_ORDER = [
    "任务闭环（专线投诉诊断）",
    "信息协商（缺参要数）",
    "目标协商（修复目标确认）",
    "可行性协商（修复方案确认）",
    "通用（终止）",
    "未分类",
]
DOMAIN_FALLBACK_TAG = {
    "information": "信息协商（缺参要数）",
    "target": "目标协商（修复目标确认）",
    "feasibility": "可行性协商（修复方案确认）",
    "abort": "通用（终止）",
    "common": "通用（终止）",
}

REVIEW_STATUSES = ("通过", "有疑问", "否决")

XLSX_HEADERS = [
    "id", "业务域", "阶段", "API动作", "输入摘要", "期望结果", "优先级", "语言", "标签",
    "评审结论", "评审意见",
]


# --- corpus loading -----------------------------------------------------------------------------------------

def json_line_of(raw_text: str, case_id: str) -> int:
    """1-based line number of the `"id": "<case_id>"` occurrence inside the corpus JSON file."""
    needle = f'"id": "{case_id}"'
    for lineno, line in enumerate(raw_text.splitlines(), start=1):
        if needle in line:
            return lineno
    return 1


def load_corpus() -> list[dict]:
    """Load all case and scenario records, enriched with source location and derived business facets."""
    records: list[dict] = []
    if not CORPUS_ROOT.is_dir():
        raise SystemExit(f"corpus root not found: {CORPUS_ROOT}")
    for family in (*CASE_FAMILIES, SCENARIO_FAMILY):
        family_dir = CORPUS_ROOT / family
        if not family_dir.is_dir():
            continue
        for path in sorted(family_dir.glob("*.json")):
            raw = path.read_text(encoding="utf-8")
            for entry in json.loads(raw):
                entry["_family"] = family
                entry["_file"] = str(path.relative_to(REPO_ROOT)).replace("\\", "/")
                entry["_line"] = json_line_of(raw, entry["id"])
                records.append(entry)
    return records


def api_action(api: str) -> str:
    return API_ACTION_ZH.get(api, api)


def business_domain(rec: dict) -> str:
    uri = rec.get("templateUri") or ""
    if not uri and rec["_family"] == SCENARIO_FAMILY:
        # scenario records carry templateUri on their steps, not at top level
        for step in rec.get("steps", []):
            uri = step.get("templateUri") or ""
            if uri:
                break
    for segment, name in TEMPLATE_URI_ZH.items():
        if segment in uri:
            return "通用（终止）" if segment == "common" else name
    tags = rec.get("tags")
    if tags is None and rec["_family"] == SCENARIO_FAMILY:
        tags = [t for step in rec.get("steps", []) for t in step.get("tags", [])]
    for tag, name in DOMAIN_FALLBACK_TAG.items():
        if tag in (tags or []):
            return name
    return "未分类"


def stage_of(rec: dict) -> str:
    if rec["_family"] == SCENARIO_FAMILY:
        return "端到端场景"
    api = rec.get("api", "")
    if api.startswith("generate"):
        return "生成"
    if api.startswith("validate"):
        return "校验"
    return "其他"


def template_uri_zh(uri: str | None) -> str:
    if not uri:
        return ""
    parts = uri.split("/")
    domain = TEMPLATE_URI_ZH.get(parts[1], parts[1]) if len(parts) > 1 else uri
    phase = PHASE_ZH.get(parts[2], parts[2]) if len(parts) > 2 else ""
    version = parts[3] if len(parts) > 3 else ""
    label = domain
    if phase:
        label += f"·{phase}"
    if version:
        label += f" {version}"
    return label


# --- technical -> business translation ----------------------------------------------------------------------

def translate_llm_script(llm: dict | None) -> str:
    """`[{"$fail":"non-json"}, {"$ref":"..."}]` -> 模拟大模型第 1 次返回非法数据、第 2 次返回正常内容..."""
    if not llm:
        return ""
    script = llm.get("script") or []
    if not script:
        return ""
    parts: list[str] = []
    for index, item in enumerate(script, start=1):
        if isinstance(item, dict) and "$fail" in item:
            mode = FAIL_MODE_ZH.get(item["$fail"], item["$fail"])
            parts.append(f"第 {index} 次返回{mode}")
        elif isinstance(item, dict) and "$ref" in item:
            parts.append(f"第 {index} 次返回正常内容（{item['$ref']}）")
        else:
            parts.append(f"第 {index} 次返回指定的原始内容")
    if len(script) == 1:
        # single-shot wording without the enumeration reads better
        text = "模拟大模型" + parts[0].replace("第 1 次", "")
    else:
        text = "模拟大模型" + "、".join(parts)
    if "maxAttempts" in llm:
        text += f"（最大尝试次数 {llm['maxAttempts']}）"
    return text


def translate_input(rec: dict) -> str:
    """Input summary in business language: natural-language text, structured data, or prompt under validation."""
    chunks: list[str] = []
    inp = rec.get("input") or {}
    text = inp.get("text")
    if isinstance(text, dict):
        zh = text.get("zh-CN") or text.get("en-US") or ""
        langs = "（双语输入）" if len(text) > 1 else ""
        chunks.append(f"自然语言{langs}：{clip(zh, 90)}")
    elif isinstance(text, str):
        chunks.append(f"自然语言：{clip(text, 90)}")
    data = inp.get("data")
    if data is not None:
        chunks.append(f"结构化数据：{describe_data(data)}")
    prompt = rec.get("prompt")
    if isinstance(prompt, dict):
        if "golden" in prompt:
            chunks.append(f"待校验消息取自 golden 基准 {prompt['golden']}")
        elif "fromStep" in prompt:
            chunks.append(f"待校验消息取自第 {prompt['fromStep']} 步的生成结果")
        elif "text" in prompt:
            chunks.append(f"待校验消息：{clip(str(prompt['text']), 90)}")
    schema = rec.get("schema")
    if isinstance(schema, dict) and "$ref" in schema:
        ref = schema["$ref"].split("/")[-1]
        chunks.append(f"参数模式：{SCHEMA_ZH.get(ref, ref)}")
    return "；".join(chunks)


def describe_data(data) -> str:
    if not isinstance(data, dict):
        return "（原始结构化数据）"
    parts: list[str] = []
    if "conclusion" in data:
        parts.append(f"结论={data['conclusion']}")
    items = data.get("items")
    if isinstance(items, list):
        parts.append(f"{len(items)} 个信息项")
    if "relationship" in data:
        parts.append(f"关系={data['relationship']}")
    if "reason" in data:
        parts.append(f"终止原因={clip(str(data['reason']), 40)}")
    return "、".join(parts) if parts else "（原始结构化数据）"


def translate_expect(rec: dict) -> str:
    """Expectation summary in business language."""
    expect = rec.get("expect") or {}
    outcome = expect.get("outcome")
    lines: list[str] = []
    if outcome == "success":
        lines.append("成功")
        golden = expect.get("promptTextEqualsGolden")
        if golden:
            lines.append(f"渲染消息与 golden 基准一致（{golden}）")
        params = expect.get("params")
        if isinstance(params, dict):
            rendered = "、".join(f"{TASK_PARAM_ZH.get(k, k)}={v}" for k, v in params.items())
            lines.append(f"提取参数：{clip(rendered, 120)}")
        missing = expect.get("missingParams")
        if isinstance(missing, list) and missing:
            rendered = "、".join(TASK_PARAM_ZH.get(k, k) for k in missing)
            lines.append(f"缺失参数（协商要数动因）：{rendered}")
    elif outcome == "failure":
        code = expect.get("code")
        lines.append(f"失败：{ERROR_CODE_ZH.get(code, code or '（无错误码）')}" + (f"（{code}）" if code else ""))
        exception = expect.get("exception")
        if exception:
            lines.append(f"异常类型：{EXCEPTION_ZH.get(exception, exception)}")
        slot_errors = expect.get("slotErrors") or []
        if slot_errors:
            rendered = "、".join(
                f"{e.get('slot')}：{ERROR_CODE_ZH.get(e.get('code'), e.get('code'))}" for e in slot_errors
            )
            lines.append(f"槽位错误：{rendered}")
        contains = expect.get("messageContains")
        if contains:
            lines.append(f"错误消息包含：{' / '.join(map(str, contains))}")
    llm_calls = expect.get("llmCalls")
    if llm_calls is not None:
        lines.append("不发起任何大模型调用" if llm_calls == 0 else f"共调用大模型 {llm_calls} 次")
    return "；".join(lines)


def translate_expect_flow(rec: dict) -> str:
    flow = rec.get("expectFlow") or {}
    if not flow:
        return ""
    parts: list[str] = []
    terminal = flow.get("terminalCondition")
    if terminal:
        parts.append(f"终态：{TERMINAL_ZH.get(terminal, terminal)}")
    if "roundsUsed" in flow:
        parts.append(f"共 {flow['roundsUsed']} 轮")
    if flow.get("distinctMessages"):
        parts.append("各轮消息互不相同")
    return "，".join(parts)


def clip(text: str, limit: int) -> str:
    text = str(text).replace("\n", " ")
    return text if len(text) <= limit else text[: limit - 1] + "…"


def languages_label(rec: dict) -> str:
    return "+".join(rec.get("languages", []))


def review_status_map() -> dict:
    if STATUS_PATH.is_file():
        try:
            data = json.loads(STATUS_PATH.read_text(encoding="utf-8"))
            return data.get("cases", {})
        except (OSError, ValueError):
            return {}
    return {}


# --- HTML rendering -----------------------------------------------------------------------------------------

CSS = """
:root { color-scheme: light; }
* { box-sizing: border-box; }
body { font-family: 'Microsoft YaHei', 'PingFang SC', 'Segoe UI', sans-serif; margin: 0; background: #f5f6f8; color: #1f2430; }
header { background: #1f3a5f; color: #fff; padding: 18px 28px; }
header h1 { margin: 0 0 4px; font-size: 20px; }
header p { margin: 0; font-size: 12px; opacity: .85; }
nav { background: #fff; border-bottom: 1px solid #e3e6ea; padding: 8px 28px; font-size: 13px; position: sticky; top: 0; }
nav a { color: #1f3a5f; text-decoration: none; margin-right: 18px; font-weight: 600; }
main { padding: 20px 28px 48px; max-width: 1180px; margin: 0 auto; }
h2 { font-size: 17px; border-left: 4px solid #1f3a5f; padding-left: 10px; margin: 34px 0 14px; }
h3 { font-size: 14px; margin: 20px 0 10px; color: #1f3a5f; }
.cards { display: flex; flex-wrap: wrap; gap: 12px; }
.kpi { background: #fff; border: 1px solid #e3e6ea; border-radius: 8px; padding: 14px 18px; min-width: 150px; flex: 1; }
.kpi .num { font-size: 26px; font-weight: 700; color: #1f3a5f; }
.kpi .lbl { font-size: 12px; color: #667; margin-top: 2px; }
table { border-collapse: collapse; background: #fff; width: 100%; font-size: 13px; margin: 8px 0 16px; }
th, td { border: 1px solid #e3e6ea; padding: 6px 10px; text-align: left; vertical-align: top; }
th { background: #eef1f5; font-weight: 600; white-space: nowrap; }
.badge { display: inline-block; border-radius: 10px; padding: 1px 9px; font-size: 11px; font-weight: 600; white-space: nowrap; }
.b-P0 { background: #fdecea; color: #b3261e; }
.b-P1 { background: #fff4e0; color: #a05a00; }
.b-P2 { background: #eef1f5; color: #556; }
.b-lang { background: #e8f0fe; color: #1a4d8f; }
.b-ok { background: #e3f4e4; color: #1c6b2b; }
.b-doubt { background: #fff4e0; color: #a05a00; }
.b-veto { background: #fdecea; color: #b3261e; }
.b-none { background: #eef1f5; color: #778; }
.b-risk { background: #fdecea; color: #b3261e; }
.case { background: #fff; border: 1px solid #e3e6ea; border-radius: 8px; padding: 12px 16px; margin-bottom: 12px; }
.case .title { font-weight: 700; font-size: 14px; margin-bottom: 6px; }
.case .meta { margin: 2px 0; font-size: 12.5px; color: #334; }
.case .meta b { color: #1f3a5f; }
.case .src { font-size: 11px; color: #99a; margin-top: 6px; }
.pingpong td { height: 100%; }
.pp-step { background: #f0f5fb; border-radius: 6px; padding: 8px 10px; margin: 4px 0; font-size: 12.5px; }
.pp-step .api { font-weight: 700; color: #1f3a5f; }
footer { padding: 16px 28px; color: #889; font-size: 11px; }
.risk { background: #fff; border: 1px solid #f0c8c4; border-left: 4px solid #b3261e; border-radius: 6px; padding: 10px 14px; margin: 8px 0; font-size: 13px; }
"""


def esc(value) -> str:
    return html.escape(str(value), quote=True)


def badge(value: str, cls: str) -> str:
    return f'<span class="badge {cls}">{esc(value)}</span>'


def status_badge(status: str) -> str:
    mapping = {"通过": "b-ok", "有疑问": "b-doubt", "否决": "b-veto"}
    if status in mapping:
        return badge("评审" + status, mapping[status])
    return badge("未评审", "b-none")


def render_html(records: list[dict], statuses: dict) -> str:
    total = len(records)
    expanded = sum(len(r.get("languages", [])) for r in records)
    scenarios = [r for r in records if r["_family"] == SCENARIO_FAMILY]
    cases = [r for r in records if r["_family"] != SCENARIO_FAMILY]
    bilingual = sum(1 for r in records if len(r.get("languages", [])) > 1)
    zh_only = sum(1 for r in records if r.get("languages") == ["zh-CN"])
    en_only = sum(1 for r in records if r.get("languages") == ["en-US"])
    priorities = {p: sum(1 for r in records if r.get("priority") == p) for p in ("P0", "P1", "P2")}
    no_priority = total - sum(priorities.values())
    reviewed = sum(1 for s in statuses.values() if s.get("status") in REVIEW_STATUSES)
    doc_gaps = [r for r in records if "doc-gap" in r.get("tags", [])]
    p2_records = [r for r in records if r.get("priority") == "P2"]

    out = io.StringIO()
    w = out.write
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    w("<!DOCTYPE html><html lang='zh-CN'><head><meta charset='utf-8'>")
    w("<meta name='viewport' content='width=device-width, initial-scale=1'>")
    w(f"<title>A2A-T 协商测试语料业务评审 · {now}</title><style>{CSS}</style></head><body>")
    w(f"<header><h1>专线投诉诊断协商测试语料业务评审材料</h1>"
      f"<p>单一业务域：传输专线业务投诉诊断五步闭环（Q23）· "
      f"角色：A=工作台（client，任务发起/补数方） / B=OMC（server，执行/要数方，协商发起方）</p>"
      f"<p>自动生成于 {now} · 语料投影，请勿手工编辑 · 生成命令：python tools/corpus_review.py export</p></header>")
    w("<nav><a href='#dashboard'>总览仪表盘</a><a href='#cases'>分域用例卡片</a>"
      "<a href='#scenarios'>场景时序表</a><a href='#notes'>评审说明</a></nav><main>")

    # ---- dashboard ----
    w("<h2 id='dashboard'>总览仪表盘</h2><div class='cards'>")
    kpis = [
        (total, "用例总数（基记录）"),
        (expanded, "双语展开后执行数"),
        (len(cases), "单步用例"),
        (len(scenarios), "端到端场景"),
        (bilingual, "双语用例"),
        (priorities["P0"], "P0 用例"),
        (reviewed, "已评审"),
    ]
    for num, label in kpis:
        w(f"<div class='kpi'><div class='num'>{num}</div><div class='lbl'>{esc(label)}</div></div>")
    w("</div>")

    w("<h3>家族分布</h3><table><tr><th>家族</th><th>数量</th></tr>")
    for family in (*CASE_FAMILIES, SCENARIO_FAMILY):
        count = sum(1 for r in records if r["_family"] == family)
        w(f"<tr><td>{esc(family)}</td><td>{count}</td></tr>")
    w("</table>")

    w("<h3>双语 parity</h3><table><tr><th>双语（zh-CN + en-US）</th><th>仅 zh-CN</th><th>仅 en-US</th></tr>"
      f"<tr><td>{bilingual}</td><td>{zh_only}</td><td>{en_only}</td></tr></table>"
      "<p style='font-size:12px;color:#556'>parity 契约由 CorpusContractTest 严格模式保障：happy 用例双语都成功、"
      "失败用例双语同错误码、双语 golden fixture 数量相同、展开后用例数相同。</p>")

    w("<h3>优先级分布</h3><table><tr><th>优先级</th><th>数量</th></tr>")
    for p in ("P0", "P1", "P2"):
        w(f"<tr><td>{badge(p, 'b-' + p)}</td><td>{priorities[p]}</td></tr>")
    if no_priority:
        w(f"<tr><td>（场景默认）</td><td>{no_priority}</td></tr>")
    w("</table>")

    w("<h3>业务域 × 阶段覆盖矩阵</h3><table><tr><th>业务域</th><th>生成</th><th>校验</th><th>端到端场景</th><th>合计</th></tr>")
    for domain in DOMAIN_ORDER:
        subset = [r for r in records if business_domain(r) == domain]
        if not subset:
            continue
        gen = sum(1 for r in subset if stage_of(r) == "生成")
        val = sum(1 for r in subset if stage_of(r) == "校验")
        e2e = sum(1 for r in subset if stage_of(r) == "端到端场景")
        w(f"<tr><td><b>{esc(domain)}</b></td><td>{gen}</td><td>{val}</td><td>{e2e}</td><td>{len(subset)}</td></tr>")
    w("</table>")

    w(f"<h3>协商错误码覆盖（{len(ERROR_CODE_ZH)} 个，CorpusContractTest 全量严格模式）</h3>"
    "<table><tr><th>错误码</th><th>业务含义</th><th>失败用例数</th></tr>")
    for code, zh in ERROR_CODE_ZH.items():
        count = sum(1 for r in records if (r.get("expect") or {}).get("code") == code)
        marker = " style='color:#b3261e;font-weight:700'" if count == 0 else ""
        w(f"<tr><td><code>{esc(code)}</code></td><td>{esc(zh)}</td><td{marker}>{count}</td></tr>")
    w("</table>")

    w("<h3>风险区</h3>")
    w(f"<div class='risk'>未评审用例：<b>{total - reviewed}</b> / {total}（评审闭环见下方评审说明）</div>")
    w(f"<div class='risk'>doc-gap 探针：<b>{len(doc_gaps)}</b> 个 —— 依据文档声明但未经验证的行为探针，"
      f"结论需与最新生产 Javadoc 对齐：" + "、".join(esc(r["id"]) for r in doc_gaps) + "</div>")
    w(f"<div class='risk'>P2 低优先级用例：<b>{len(p2_records)}</b> 个 —— "
      + "、".join(esc(r["id"]) for r in p2_records) + "</div>")

    # ---- per-domain case cards ----
    w("<h2 id='cases'>分域用例卡片</h2>")
    for domain in DOMAIN_ORDER:
        subset = [r for r in records if r["_family"] != SCENARIO_FAMILY and business_domain(r) == domain]
        if not subset:
            continue
        w(f"<h3>{esc(domain)}（{len(subset)} 个用例）</h3>")
        for r in subset:
            status = statuses.get(r["id"], {}).get("status", "")
            w("<div class='case'>")
            w(f"<div class='title'>{esc(r['id'])} · {esc(r.get('summary', ''))}</div>")
            w(f"<div class='meta'><b>动作</b>：{esc(api_action(r.get('api', '')))}"
              f"　<b>模板</b>：{esc(template_uri_zh(r.get('templateUri')))}</div>")
            w(f"<div class='meta'><b>输入</b>：{esc(translate_input(r))}</div>")
            llm_desc = translate_llm_script(r.get("llm"))
            if llm_desc:
                w(f"<div class='meta'><b>大模型</b>：{esc(llm_desc)}</div>")
            inject = INJECT_ZH.get(r.get("inject", ""))
            if inject:
                w(f"<div class='meta'><b>故障注入</b>：{esc(inject)}</div>")
            w(f"<div class='meta'><b>期望</b>：{esc(translate_expect(r))}</div>")
            w("<div style='margin-top:6px'>")
            w(badge(r.get("priority", "P2"), "b-" + r.get("priority", "P2")))
            w(" " + badge(languages_label(r), "b-lang"))
            w(" " + status_badge(status))
            if "doc-gap" in r.get("tags", []):
                w(" " + badge("doc-gap 探针", "b-risk"))
            w("</div>")
            w(f"<div class='src'>来源：{esc(r['_file'])} 第 {r['_line']} 行 · 标签：{esc(', '.join(r.get('tags', [])))}</div>")
            w("</div>")

    # ---- scenario ping-pong tables ----
    w("<h2 id='scenarios'>场景时序表（A/B 乒乓）</h2>")
    if not scenarios:
        w("<p>（无场景记录）</p>")
    for r in scenarios:
        status = statuses.get(r["id"], {}).get("status", "")
        w(f"<h3>{esc(r['id'])} · {esc(r.get('summary', ''))} " + status_badge(status) + "</h3>")
        w("<table class='pingpong'><tr><th style='width:52px'>步骤</th><th>A（工作台，发起/补数方）</th><th>B（OMC，执行/要数方）</th></tr>")
        for step in r.get("steps", []):
            role = step.get("role", "A")
            col = "B" if role == "B" else "A"
            detail = io.StringIO()
            detail.write(f"<div class='pp-step'><div class='api'>{esc(api_action(step.get('api', '')))}</div>")
            inp = translate_input(step)
            if inp:
                detail.write(f"<div>输入：{esc(inp)}</div>")
            llm_desc = translate_llm_script(step.get("llm"))
            if llm_desc:
                detail.write(f"<div>{esc(llm_desc)}</div>")
            detail.write(f"<div>期望：{esc(translate_expect(step))}</div></div>")
            cells = {"A": "<td></td>", "B": "<td></td>"}
            cells[col] = f"<td>{detail.getvalue()}</td>"
            w(f"<tr><td style='text-align:center'>{step.get('step', '')}</td>{cells['A']}{cells['B']}</tr>")
        w("</table>")
        flow = translate_expect_flow(r)
        if flow:
            w(f"<p style='font-size:12.5px'><b>流程期望</b>：{esc(flow)}</p>")
        w(f"<div class='src'>来源：{esc(r['_file'])} 第 {r['_line']} 行</div>")

    # ---- notes ----
    w("<h2 id='notes'>评审说明（闭环流程）</h2>")
    w("<ol style='font-size:13px;line-height:1.9'>"
      "<li>本材料由 <code>tools/corpus_review.py export</code> 从语料自动生成，是语料的只读投影，请勿手工编辑。</li>"
      "<li>业务专家在 <code>corpus-review.xlsx</code> 最后两列批注：评审结论（通过 / 有疑问 / 否决）与评审意见。</li>"
      "<li>运行 <code>python tools/corpus_review.py collect --marked &lt;批注后的xlsx&gt;</code> 回读标记，"
      "生成 findings 报告与 review-status.json，并精确映射回 JSON 源文件与行号。</li>"
      "<li>有疑问 / 否决的用例驱动语料修改，随后重新 export 进入下一轮评审。</li></ol>")
    w("</main><footer>A2A-T negotiation test corpus review projection · tools/corpus_review.py</footer></body></html>")
    return out.getvalue()


# --- xlsx rendering ------------------------------------------------------------------------------------------

def case_row(rec: dict) -> list[str]:
    tags = ", ".join(rec.get("tags", []))
    if rec["_family"] == SCENARIO_FAMILY:
        apis = sorted({s.get("api", "") for s in rec.get("steps", [])})
        api_label = " → ".join(api_action(a) for a in apis)
        input_label = f"端到端流程（{len(rec.get('steps', []))} 步）：" + clip(rec.get("summary", ""), 60)
        expect_label = translate_expect_flow(rec) or "见语料 expectFlow"
    else:
        api_label = api_action(rec.get("api", ""))
        input_label = translate_input(rec)
        expect_label = translate_expect(rec)
    return [
        rec["id"], business_domain(rec), stage_of(rec), api_label, input_label, expect_label,
        rec.get("priority", ""), languages_label(rec), tags, "", "",
    ]


def write_xlsx(rows: list[list[str]]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError:
        print("[warn] openpyxl 未安装（pip install openpyxl），跳过 xlsx 输出，仅生成 HTML 与 CSV", file=sys.stderr)
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "corpus-review"
    header_fill = PatternFill("solid", fgColor="1F3A5F")
    for col, value in enumerate(XLSX_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=value)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_index, row in enumerate(rows, start=2):
        for col, value in enumerate(row, start=1):
            ws.cell(row=row_index, column=col, value=value).alignment = Alignment(vertical="top", wrap_text=True)
    widths = [14, 12, 11, 26, 46, 46, 8, 14, 30, 12, 30]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    validation = DataValidation(type="list", formula1='"通过,有疑问,否决"', allow_blank=True)
    ws.add_data_validation(validation)
    validation.add(f"J2:J{len(rows) + 1}")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(XLSX_HEADERS))}{len(rows) + 1}"
    REVIEW_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(XLSX_PATH)
    print(f"[ok] wrote {XLSX_PATH.relative_to(REPO_ROOT)} ({len(rows)} rows)")


def write_csv(rows: list[list[str]]) -> None:
    REVIEW_DIR.mkdir(parents=True, exist_ok=True)
    with CSV_PATH.open("w", encoding="utf-8-sig", newline="") as fh:
        csv.writer(fh).writerows([XLSX_HEADERS, *rows])
    print(f"[ok] wrote {CSV_PATH.relative_to(REPO_ROOT)} ({len(rows)} rows, degraded/backup format)")


# --- export / collect commands -------------------------------------------------------------------------------

def cmd_export() -> int:
    records = load_corpus()
    statuses = review_status_map()
    rows = [case_row(r) for r in records]
    REVIEW_DIR.mkdir(parents=True, exist_ok=True)
    HTML_PATH.write_text(render_html(records, statuses), encoding="utf-8")
    print(f"[ok] wrote {HTML_PATH.relative_to(REPO_ROOT)} "
          f"({len(records)} records, {sum(len(r.get('languages', [])) for r in records)} expanded)")
    write_xlsx(rows)
    write_csv(rows)
    return 0


def read_marked(path: Path) -> list[dict]:
    """Read annotated rows from xlsx (openpyxl) or csv; each row keeps id / 评审结论 / 评审意见."""
    rows: list[dict] = []
    if path.suffix.lower() == ".csv":
        with path.open(encoding="utf-8-sig", newline="") as fh:
            for raw in csv.DictReader(fh):
                rows.append({
                    "id": (raw.get("id") or "").strip(),
                    "conclusion": (raw.get("评审结论") or "").strip(),
                    "comment": (raw.get("评审意见") or "").strip(),
                })
        return rows
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit("[error] 回读 xlsx 需要 openpyxl（pip install openpyxl）；或将批注结果另存为 CSV 后重试")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(it, [])]
    try:
        id_col = header.index("id")
        concl_col = header.index("评审结论")
        comment_col = header.index("评审意见")
    except ValueError as exc:
        raise SystemExit(f"[error] 批注文件表头缺少必需列: {exc}")
    for values in it:
        if values is None or id_col >= len(values):
            continue
        case_id = values[id_col]
        if case_id is None or str(case_id).strip() == "":
            continue
        def cell(index: int) -> str:
            return str(values[index]).strip() if index < len(values) and values[index] is not None else ""
        rows.append({"id": str(case_id).strip(), "conclusion": cell(concl_col), "comment": cell(comment_col)})
    wb.close()
    return rows


def normalize_conclusion(value: str) -> str | None:
    lowered = value.strip().lower()
    mapping = {"通过": "通过", "ok": "通过", "pass": "通过", "yes": "通过", "y": "通过",
               "有疑问": "有疑问", "疑问": "有疑问", "question": "有疑问", "q": "有疑问",
               "否决": "否决", "reject": "否决", "no": "否决", "n": "否决"}
    return mapping.get(lowered)


def cmd_collect(marked: Path) -> int:
    if not marked.is_file():
        raise SystemExit(f"[error] marked file not found: {marked}")
    records = {r["id"]: r for r in load_corpus()}
    marked_rows = read_marked(marked)
    statuses: dict[str, dict] = {}
    problems: list[str] = []
    for row in marked_rows:
        conclusion = normalize_conclusion(row["conclusion"])
        if conclusion is None:
            if row["conclusion"]:
                problems.append(f"无法识别的评审结论 '{row['conclusion']}'（用例 {row['id']}），应为 通过/有疑问/否决")
            continue
        if row["id"] not in records:
            problems.append(f"批注中的用例 id {row['id']} 不存在于语料")
            continue
        rec = records[row["id"]]
        statuses[row["id"]] = {
            "status": conclusion,
            "comment": row["comment"],
            "source": {"file": rec["_file"], "line": rec["_line"]},
        }

    counts = {s: sum(1 for v in statuses.values() if v["status"] == s) for s in REVIEW_STATUSES}
    now = datetime.now().isoformat(timespec="seconds")
    status_doc = {
        "generatedAt": now,
        "markedFrom": str(marked).replace("\\", "/"),
        "summary": {
            "total": len(records),
            "reviewed": len(statuses),
            "pending": len(records) - len(statuses),
            **counts,
        },
        "cases": statuses,
    }
    REVIEW_DIR.mkdir(parents=True, exist_ok=True)
    STATUS_PATH.write_text(json.dumps(status_doc, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[ok] wrote {STATUS_PATH.relative_to(REPO_ROOT)} "
          f"(reviewed={len(statuses)}/{len(records)}, 通过={counts['通过']}, 有疑问={counts['有疑问']}, 否决={counts['否决']})")

    lines = [
        "# 协商测试语料评审 findings",
        "",
        f"- 回读时间：{now}",
        f"- 批注文件：`{str(marked).replace(chr(92), '/')}`",
        f"- 评审进度：{len(statuses)} / {len(records)}（通过 {counts['通过']}、有疑问 {counts['有疑问']}、否决 {counts['否决']}、未评审 {len(records) - len(statuses)}）",
        "",
    ]
    if problems:
        lines += ["## 回读告警", ""]
        lines += [f"- {p}" for p in problems]
        lines.append("")
    for status, heading in (("否决", "否决（必须修改语料或期望）"), ("有疑问", "有疑问（需业务澄清）")):
        subset = [rec for cid, rec in records.items() if statuses.get(cid, {}).get("status") == status]
        lines += [f"## {heading}（{len(subset)}）", ""]
        if not subset:
            lines.append("（无）")
        for rec in subset:
            mark = statuses[rec["id"]]
            lines.append(f"### {rec['id']} — {rec.get('summary', '')}")
            lines.append(f"- 位置：`{mark['source']['file']}` 第 {mark['source']['line']} 行")
            lines.append(f"- 意见：{mark['comment'] or '（未填写）'}")
            lines.append(f"- 业务域：{business_domain(rec)} · 阶段：{stage_of(rec)} · 优先级：{rec.get('priority', '—')}")
            lines.append(f"- 输入：{translate_input(rec)}")
            lines.append(f"- 期望：{translate_expect(rec) if rec['_family'] != SCENARIO_FAMILY else translate_expect_flow(rec)}")
            lines.append("")
    passed = [cid for cid, v in statuses.items() if v["status"] == "通过"]
    lines += [f"## 通过（{len(passed)}）", "", "、".join(passed) if passed else "（无）", ""]
    FINDINGS_PATH.write_text("\n".join(lines), encoding="utf-8")
    print(f"[ok] wrote {FINDINGS_PATH.relative_to(REPO_ROOT)}")
    for problem in problems:
        print(f"[warn] {problem}", file=sys.stderr)
    return 0


def main(argv: list[str] | None = None) -> int:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    sub = parser.add_subparsers(dest="command", required=True)
    sub.add_parser("export", help="生成 docs-local/review/corpus-review.html 与 .xlsx（评审材料投影）")
    collect = sub.add_parser("collect", help="回读批注后的 xlsx/csv，生成 findings 报告与 review-status.json")
    collect.add_argument("--marked", required=True, type=Path, help="批注后的 corpus-review xlsx 或 csv 文件")
    args = parser.parse_args(argv)
    if args.command == "export":
        return cmd_export()
    return cmd_collect(args.marked)


if __name__ == "__main__":
    raise SystemExit(main())
